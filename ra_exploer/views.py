from io import BytesIO

from django.views.generic.edit import UpdateView
from django.views.generic import View
from django.http.response import HttpResponseBadRequest, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls.base import reverse
from django.http import HttpResponse

import numpy as np
import plotly.express as px
import pandas as pd
from plotly.offline import plot
import csv
from openpyxl.reader.excel import load_workbook

# Create your views here.

from .forms import SearchFrom, UploadFileForm
from .models import VHR, DeltaAngle, LiquidCrystal, LowTemperatureOperation, Polyimide, PressureCookingTest, Seal, SealWVTR, Validator, Vender, File
from .models import Adhesion, LowTemperatureStorage
from utils.TR2_tools import tr2_score


def index(request):
    """View function for home page of site."""
    clean = request.GET.get('clean')
    if clean:
        if 'filtered LC list' in request.session:
            del request.session['filtered LC list']

    query = {
        'LC': [x[0] for x in LiquidCrystal.objects.all().values_list('name')],
        'PI': [x[0] for x in Polyimide.objects.all().values_list('name')],
        'Seal': [x[0] for x in Seal.objects.all().values_list('name')],
    }

    LCs = pd.DataFrame.from_records(
        LiquidCrystal.objects.all().order_by('name').values('name'))['name'].to_list()
    if 'filtered LC list' in request.session:
        LCs = pd.read_json(request.session['filtered LC list'])['LC'].to_list() + ['N.A.']
    PIs = Polyimide.objects.all().order_by('name')
    seals = Seal.objects.all().order_by('name')

    valid_adhesion = Validator.objects.get_or_create(name='adhesion test')[0]
    valid_LTO = Validator.objects.get_or_create(name='LTO')[0]
    valid_LTS = Validator.objects.get_or_create(name='LTS')[0]
    valid_delta_angle = Validator.objects.get_or_create(name='Δ angle')[0]
    valid_VHR = Validator.objects.get_or_create(name='VHR(heat)')[0]
    valid_PCT = Validator.objects.get_or_create(name='PCT')[0]
    valid_SealWVTR = Validator.objects.get_or_create(name='Seal WVTR')[0]

    # Render the HTML template index.html with the data in the context variable
    return render(
        request,
        'index.html',
        context=(
            {
                'query': query,
                'LCs': LCs,
                'PIs': PIs,
                'seals': seals,
                'valid_adhesion': valid_adhesion,
                'valid_LTO': valid_LTO,
                'valid_LTS': valid_LTS,
                'valid_delta_angle': valid_delta_angle,
                'valid_VHR': valid_VHR,
                'valid_PCT': valid_PCT,
                'valid_SealWVTR': valid_SealWVTR,
                'opt_lc_list': request.session.get('filtered LC list')
            }
        )
    )


def import_adhesion(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            # print('valid!')
            ws = load_workbook(
                filename=request.FILES['file'].file).worksheets[0]
            header = True  # using for skip header, it should have a nicer method?

            for row in ws.values:
                if header:
                    header = False
                    continue
                lc, _ = LiquidCrystal.objects.get_or_create(name=row[1])
                pi, _ = Polyimide.objects.get_or_create(name=row[2])
                seal, _ = Seal.objects.get_or_create(name=row[3])
                vender, _ = Vender.objects.get_or_create(name=row[8])
                file, _ = File.objects.get_or_create(name=row[9])
                if not Adhesion.objects.filter(LC=lc, PI=pi, seal=seal, vender=vender, file_source=file,
                                               adhesion_interface=row[5], method=row[6]).exists():
                    Adhesion.objects.create(LC=lc, PI=pi, seal=seal, vender=vender, file_source=file,
                                            adhesion_interface=row[5], method=row[6], value=row[4], peeling=row[7])
            return redirect(reverse('index'))

        else:
            return HttpResponseBadRequest()
    return redirect(reverse('index'))


def import_LTO(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            print('valid!')
            ws = load_workbook(
                filename=request.FILES['file'].file).worksheets[0]

            header = True  # using for skip header, it should have a nicer method?
            for row in ws.values:
                if header:
                    header = False
                    continue
                LC, _ = LiquidCrystal.objects.get_or_create(name=row[1])
                PI, _ = Polyimide.objects.get_or_create(name=row[2])
                seal, _ = Seal.objects.get_or_create(name=row[3])
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                JarTestSeal, _ = Seal.objects.get_or_create(name=row[7])
                if not LowTemperatureOperation.objects.filter(LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                                              storage_condition=row[5], SLV_condition=row[6],
                                                              JarTestSeal=JarTestSeal,
                                                              measure_temperature=row[8]).exists():

                    LowTemperatureOperation.objects.create(LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                                           storage_condition=row[5], SLV_condition=row[6],
                                                           JarTestSeal=JarTestSeal, measure_temperature=row[8],
                                                           value=LowTemperatureOperation.value_mapping[row[4]])
            return redirect(reverse('index'))

        else:
            return HttpResponseBadRequest()
    return redirect(reverse('index'))


def import_LTS(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            print('valid!')
            ws = load_workbook(
                filename=request.FILES['file'].file).worksheets[0]

            header = True  # using for skip header, it should have a nicer method?
            for row in ws.values:
                if header:
                    header = False
                    continue
                LC, _ = LiquidCrystal.objects.get_or_create(name=row[1])
                PI, _ = Polyimide.objects.get_or_create(name=row[2])
                seal, _ = Seal.objects.get_or_create(name=row[3])
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                JarTestSeal, _ = Seal.objects.get_or_create(name=row[7])
                if not LowTemperatureStorage.objects.filter(LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                                            storage_condition=row[5], SLV_condition=row[6],
                                                            JarTestSeal=JarTestSeal,
                                                            measure_temperature=row[8]).exists():
                    LowTemperatureStorage.objects.create(LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                                         storage_condition=row[5], SLV_condition=row[6],
                                                         JarTestSeal=JarTestSeal, measure_temperature=row[8],
                                                         value=row[4])
            return redirect(reverse('index'))

        else:
            return HttpResponseBadRequest()
    return redirect(reverse('index'))


def import_DeltaAngle(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            print('valid!')
            ws = load_workbook(
                filename=request.FILES['file'].file).worksheets[0]

            header = True  # using for skip header, it should have a nicer method?
            for row in ws.values:
                if header:
                    header = False
                    continue
                LC, _ = LiquidCrystal.objects.get_or_create(name=row[1])
                PI, _ = Polyimide.objects.get_or_create(name=row[2])
                seal, _ = Seal.objects.get_or_create(name=row[3])
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                if not DeltaAngle.objects.filter(LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                                 measure_voltage=row[5], measure_freq=row[6], measure_time=row[7],
                                                 measure_temperature=row[8]).exists():
                    DeltaAngle.objects.create(value=row[4], LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                              measure_voltage=row[5], measure_freq=row[6], measure_time=row[7],
                                              measure_temperature=row[8])
            return redirect(reverse('index'))

        else:
            return HttpResponseBadRequest()
    return redirect(reverse('index'))


def import_VHR(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            print('valid!')
            ws = load_workbook(
                filename=request.FILES['file'].file).worksheets[0]

            header = True  # using for skip header, it should have a nicer method?
            for row in ws.values:
                if header:
                    header = False
                    continue
                LC, _ = LiquidCrystal.objects.get_or_create(name=row[1])
                PI, _ = Polyimide.objects.get_or_create(name=row[2])
                seal, _ = Seal.objects.get_or_create(name=row[3])
                vender, _ = Vender.objects.get_or_create(name=row[9])
                file, _ = File.objects.get_or_create(name=row[10])
                if not VHR.objects.filter(LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                          measure_voltage=row[5], measure_freq=row[6], measure_temperature=row[7],
                                          UV_aging=row[8]).exists():
                    VHR.objects.create(value=row[4], LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                       measure_voltage=row[5], measure_freq=row[6], measure_temperature=row[7],
                                       UV_aging=row[8])
            return redirect(reverse('index'))

        else:
            return HttpResponseBadRequest()
    return redirect(reverse('index'))


def import_PCT(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            ws = load_workbook(
                filename=request.FILES['file'].file
            ).worksheets[0]

            header = True
            for row in ws.values:
                if header:
                    header = False
                    continue
                LC, _ = LiquidCrystal.objects.get_or_create(name=row[1])
                PI, _ = Polyimide.objects.get_or_create(name=row[2])
                seal, _ = Seal.objects.get_or_create(name=row[3])
                vender, _ = Vender.objects.get_or_create(name=row[7])
                file, _ = File.objects.get_or_create(name=row[8])
                if not PressureCookingTest.objects.filter(LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                                          measure_condition=row[5], test_vehical=row[6],
                                                          ).exists():
                    PressureCookingTest.objects.create(value=row[4], LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                                       measure_condition=row[5], test_vehical=row[6],
                                                       )
            return redirect(reverse('index'))

        else:
            return HttpResponseBadRequest()
    return redirect(reverse('index'))


def import_SealWVTR(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            print('valid!')
            ws = load_workbook(
                filename=request.FILES['file'].file).worksheets[0]

            header = True  # using for skip header, it should have a nicer method?
            for row in ws.values:
                if header:
                    header = False
                    continue
                LC, _ = LiquidCrystal.objects.get_or_create(name=row[1])
                PI, _ = Polyimide.objects.get_or_create(name=row[2])
                seal, _ = Seal.objects.get_or_create(name=row[3])
                vender, _ = Vender.objects.get_or_create(name=row[7])
                file, _ = File.objects.get_or_create(name=row[8])
                if not SealWVTR.objects.filter(LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                               measure_condition_1=row[5], measure_condition_2=row[6],
                                               ).exists():
                    SealWVTR.objects.create(value=row[4], LC=LC, PI=PI, seal=seal, vender=vender, file_source=file,
                                            measure_condition_1=row[5], measure_condition_2=row[6],
                                            )
            return redirect(reverse('index'))

        else:
            return HttpResponseBadRequest()
    return redirect(reverse('index'))


def query_table(query, model, writer, valid=True, cmp='gt'):

    valid_val = Validator.objects.get_or_create(name=model.name)[0].value

    if 'ALL' in query['LC']:
        query['LC'] = LiquidCrystal.objects.all().values_list('name')
    if 'ALL' in query['PI']:
        query['PI'] = Polyimide.objects.all().values_list('name')
    if 'ALL' in query['Seal']:
        query['Seal'] = Seal.objects.all().values_list('name')
    if cmp == 'gt':
        results = model.objects.filter(
            LC__name__in=query['LC'],
            PI__name__in=query['PI'],
            seal__name__in=query['Seal'],
            value__gt=valid_val,
        )
    elif cmp == 'lt':
        results = model.objects.filter(
            LC__name__in=query['LC'],
            PI__name__in=query['PI'],
            seal__name__in=query['Seal'],
            value__lt=valid_val,
        )

    for result in results:

        row = ['N.A.'] * 10
        if not (result.name is None):
            row[0] = result.name
        if not (result.LC is None):
            row[1] = result.LC.name
        if not (result.PI is None):
            row[2] = result.PI.name
        if not (result.seal is None):
            row[3] = result.seal.name
        if not (result.value is None):
            if model.name == 'LTO':
                value = result.get_value_display()
            else:
                value = result.value
            row[4] = value
        if not (result.unit is None):
            row[5] = result.unit
        if not (result.value_remark is None):
            row[6] = result.value_remark()
        if not (result.vender is None):
            row[7] = result.vender.name
        if not (result.cond is None):
            row[8] = result.cond()
        if not (result.file_source is None):
            row[9] = result.file_source.name

        writer.writerow(row)
    print('query finished')


def export_results_csv(request):
    if request.method == 'POST':
        form = SearchFrom(request.POST)
        if form.is_valid():
            query = {
                'LC': form.cleaned_data['LC'],
                'PI': form.cleaned_data['PI'],
                'Seal': form.cleaned_data['Seal'],
            }
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = 'attachment; filename="results.csv"'
            writer = csv.writer(response)
            writer.writerow(['Item', 'LC', 'PI', 'Seal',
                             'Value', 'Unit', 'Value Remark', 'Vendor', 'Condition', 'file'])

            # need fixed?
            query_table(query, VHR, writer)
            query_table(query, DeltaAngle, writer, cmp='lt')
            query_table(query, Adhesion, writer)
            query_table(query, LowTemperatureOperation, writer)
            query_table(query, LowTemperatureStorage, writer)
            # query_table(query, AC_IS, writer)

            return response
    return redirect(reverse('index'))


class ValidatorUpdateView(UpdateView):
    template_name = 'validator_update.html'
    template_name_suffix = 'update'
    model = Validator
    slug_field = 'name'
    fields = ['value', 'venders', 'weight']

    def get_success_url(self):
        return reverse('index')


def filterQuery(query, model, cmp='gt'):
    validator = Validator.objects.get_or_create(name=model.name)
    valid_val = validator[0].value
    valid_venders = validator[0].venders.all()
    weight = float(validator[0].weight)

    if 'ALL' in query['LC']:
        query['LC'] = list(LiquidCrystal.objects.all().values_list('name', flat=True))
    if 'ALL' in query['PI']:
        query['PI'] = list(Polyimide.objects.all().values_list('name', flat=True))
    if 'ALL' in query['Seal']:
        query['Seal'] = list(Seal.objects.all().values_list('name', flat=True))
    if cmp == 'gt':
        result = model.objects.filter(
            LC__name__in=query['LC'],
            PI__name__in=query['PI'],
            seal__name__in=query['Seal'],
            value__gt=valid_val,
            vender__in=valid_venders,
        ).order_by('vender__name', '-value')
    elif cmp == 'lt':
        result = model.objects.filter(
            LC__name__in=query['LC'],
            PI__name__in=query['PI'],
            seal__name__in=query['Seal'],
            value__lt=valid_val,
            vender__in=valid_venders,
        ).order_by('vender__name', 'value')
    else:
        result = model.objects.all()
    result_df = pd.DataFrame.from_records(
        result.values(
            "LC__name",
            "PI__name",
            "seal__name",
            "value",
            "vender__name",
            "file_source__name",
        )
    ).rename(
        columns={
            "LC__name": "LC",
            "PI__name": "PI",
            "seal__name": "Seal",
            "vender__name": "Vender",
            "file_source__name": "file source"
        }
    )
    # Values generated by customized methods
    value_remark = []
    condition = []
    for item in result:
        value_remark += [item.value_remark()]
        condition += [item.cond()]
    if len(result_df.columns) == 6:
        result_df.insert(4, 'condition', condition)
        result_df.insert(4, 'value remark', value_remark)
        result_df.insert(0, 'item', model.name)
        result_df = result_df.astype({'value': float})

        result_mean_df = result_df.groupby(by=['LC', 'PI', 'Seal', 'Vender'], as_index=False).mean(
        ).sort_values(by=['value'], ascending=False)
        result_mean_df['configuration'] = result_mean_df['LC'] + \
            ' ' + result_mean_df['PI'] + ' ' + result_mean_df['Seal']

        def ra_formatter(x):
            return np.round(9 * x) + 1

        if model.name in ['Δ angle', 'Seal WVTR']:
            result_mean_df['score'] = tr2_score(
                result_mean_df['value'],
                cmp='lt',
                method='min-max',
                formatter=ra_formatter,
                scale=weight
            )
        else:
            result_mean_df['score'] = tr2_score(
                result_mean_df['value'],
                cmp='gt',
                method='min-max',
                formatter=ra_formatter,
                scale=weight
            )
        result_mean_df.insert(0, 'item', model.name)
        if model.name == 'LTO':
            values = []
            for item in result:
                values += [item.get_value_display()]
            result_df['value'] = values
        return result_df, result_mean_df

    return pd.DataFrame(), pd.DataFrame()


def filteredResultView(request):
    if request.method == 'POST':
        form = SearchFrom(request.POST)
        if form.is_valid():
            query = {
                'LC': form.cleaned_data['LC'],
                'PI': form.cleaned_data['PI'],
                'Seal': form.cleaned_data['Seal'],
            }
            # vhr
            vhr_df, vhr_mean_df = filterQuery(query, VHR)
            plot_vhr = None
            if len(vhr_mean_df) > 0:
                vhr_fig = px.bar(
                    vhr_mean_df,
                    x='configuration',
                    y='value',
                    color='Vender',
                    barmode='group',
                    labels={
                        'value': 'VHR(%)'
                    }
                )
                plot_vhr = plot(vhr_fig, output_type='div')

            # adhesion
            adhesion_df, adhesion_mean_df = filterQuery(query, Adhesion)
            plot_adhesion = None
            if len(adhesion_mean_df) > 0:
                adhesion_fig = px.bar(
                    adhesion_mean_df,
                    x='configuration',
                    y='value',
                    color='Vender',
                    barmode='group',
                    labels={
                        'value': 'adhesion(kgw)'
                    }
                )
                plot_adhesion = plot(adhesion_fig, output_type='div')

            # lts
            lts_df, lts_mean_df = filterQuery(query, LowTemperatureStorage)
            plot_lts = None
            if len(lts_mean_df) > 0:
                lts_fig = px.bar(
                    lts_mean_df,
                    x='configuration',
                    y='value',
                    color='Vender',
                    barmode='group',
                    labels={
                        'value': 'LTS (days)'
                    }
                )
                plot_lts = plot(lts_fig, output_type='div')

            # lto
            lto_df, _ = filterQuery(query, LowTemperatureOperation)

            plot_delta_angle = None
            delta_angle_df, delta_angle_mean_df = filterQuery(
                query, DeltaAngle, 'lt')
            if len(delta_angle_mean_df) > 0:
                delta_angle_fig = px.bar(
                    delta_angle_mean_df,
                    x='configuration',
                    y='value',
                    color='Vender',
                    barmode='group',
                    labels={
                        'value': 'Δ angle(°)'
                    }
                )
                plot_delta_angle = plot(delta_angle_fig, output_type='div')

            pct_df, pct_mean_df = filterQuery(query, PressureCookingTest)
            plot_pct = None
            if len(pct_mean_df) > 0:
                pct_fig = px.bar(
                    pct_mean_df,
                    x='configuration',
                    y='value',
                    color='Vender',
                    barmode='group',
                    labels={
                        'value': 'PCT(hours)'
                    }
                )
                plot_pct = plot(pct_fig, output_type='div')

            sealwvtr_df, sealwvtr_mean_df = filterQuery(query, SealWVTR)
            plot_sealwvtr = None
            if len(sealwvtr_mean_df) > 0:
                sealwvtr_fig = px.bar(
                    sealwvtr_mean_df,
                    x='configuration',
                    y='value',
                    color='Vender',
                    barmode='group',
                    labels={
                        'value': 'Seal WVTR'
                    }
                )
                plot_sealwvtr = plot(sealwvtr_fig, output_type='div')

            ra_result = pd.concat(
                [
                    vhr_df,
                    adhesion_df,
                    lts_df,
                    lto_df,
                    delta_angle_df,
                    pct_df,
                    sealwvtr_df,
                ],
                ignore_index=True
            )

            # ra_mean = pd.concat(
            #     [
            #         vhr_mean_df,
            #         adhesion_mean_df,
            #         lts_mean_df,
            #         delta_angle_mean_df,
            #         pct_mean_df,
            #         sealwvtr_mean_df,
            #     ],
            #     ignore_index=True
            # )

            # ra_score = ra_mean[['item', 'configuration', 'score']].groupby(
            #     by=['item', 'configuration']).mean()
            # ra_score = ra_score.unstack(level=0)
            # ra_score = ra_score.fillna(0)
            # ra_score[('score', 'Sum')] = ra_score.sum(axis=1)
            # ra_score = ra_score.sort_values(
            #     by=('score', 'Sum'), ascending=False)
            # ra_score = ra_score.droplevel(0, axis=1).reset_index()
            # ra_score.columns.name = None
            # if 'ALL' in query['LC']:
            #     query['LC'] = list(LiquidCrystal.objects.all().values_list('name', flat=True))
            # if 'ALL' in query['PI']:
            #     query['PI'] = list(Polyimide.objects.all().values_list('name', flat=True))
            # if 'ALL' in query['Seal']:
            #     query['Seal'] = list(Seal.objects.all().values_list('name', flat=True))
            print(query)
            df_LC = pd.DataFrame({'LC': query['LC']})
            df_PI = pd.DataFrame({'PI': query['PI']})
            df_seal = pd.DataFrame({'Seal': query['Seal']})

            df_LC_PI = df_LC.merge(df_PI, how='cross')
            df = df_LC_PI.merge(df_seal, how='cross')

            def is_worth_df(df):
                return not(df.empty or (df['score'].sum() == 0))

            if is_worth_df(adhesion_mean_df):
                adhesion_mean_df = adhesion_mean_df.groupby(by=['PI', 'Seal'], as_index=False).mean()
                df = df.merge(adhesion_mean_df[['PI', 'Seal', 'score']], on=['PI', 'Seal'], how='left').rename(columns={'score': 'Adhesion'})

            if is_worth_df(delta_angle_mean_df):
                delta_angle_mean_df = delta_angle_mean_df.groupby(by=['LC', 'PI'], as_index=False).mean()
                df = df.merge(delta_angle_mean_df[['LC', 'PI', 'score']], on=['LC', 'PI'], how='left').rename(columns={'score': 'Δ angle'})

            if is_worth_df(vhr_mean_df):
                vhr_mean_df = vhr_mean_df.groupby(by=['LC', 'PI', 'Seal'], as_index=False).mean()
                df = df.merge(vhr_mean_df[['LC', 'PI', 'Seal', 'score']], on=['LC', 'PI', 'Seal'], how='left').rename(columns={'score': 'VHR'})

            if is_worth_df(lts_mean_df):
                lts_mean_df = lts_mean_df.groupby(by=['LC'], as_index=False).mean()
                df = df.merge(lts_mean_df[['LC', 'score']], on=['LC'], how='left').rename(columns={'score': 'LTS'})

            if is_worth_df(pct_mean_df):
                pct_mean_df = pct_mean_df.groupby(by=['LC', 'PI', 'Seal'], as_index=False).mean()
                df = df.merge(pct_mean_df[['LC', 'PI', 'Seal', 'score']], on=['LC', 'PI', 'Seal'], how='left').rename(columns={'score': 'PCT'})

            if is_worth_df(sealwvtr_mean_df):
                sealwvtr_mean_df = sealwvtr_mean_df.groupby(by=['Seal'], as_index=False).mean()
                df = df.merge(sealwvtr_mean_df[['Seal', 'score']], on=['Seal'], how='left').rename(columns={'score': 'Seal WVTR'})
            
            df_fill0 = df.fillna(0)
            df_fill0 = df_fill0[
                (df_fill0.LC != 'N.A.') &
                (df_fill0.PI != 'N.A.') &
                (df_fill0.Seal != 'N.A.')
            ]

            item_columns = list(df.columns[3:])
            df_fill0['Sum'] = df_fill0[item_columns].sum(axis=1)
            df_fill0 = df_fill0.sort_values(by='Sum', ascending=False)
            # print(df_fill0['LC'])
            # print(df_fill0['PI'])
            # print(df_fill0['Seal'])

            df_fill0['Configuration'] = df_fill0['LC'] + ' ' + df_fill0['PI'] + ' ' + df_fill0['Seal']
            ra_score = df_fill0
            ra_score_table = ra_score.to_html(
                float_format=lambda x: f'{x:.0f}',
                classes=['table', 'table-hover', 'text-center  '],
                justify='center',
                index=False,
            )
            
            ra_plot_df = df_fill0[:10] \
                .set_index('Configuration')[item_columns + ['Sum']] \
                .unstack() \
                .reset_index() \
                .rename(columns={'level_0': 'Item', 0: 'Score'})

            ra_fig = px.bar(
                ra_plot_df,
                x='Item',
                y='Score',
                color='Configuration',
                barmode='group'
            )
            ra_fig_div = plot(ra_fig, output_type='div')

            request.session['ra_result'] = ra_result.to_json()
            request.session['ra_score'] = ra_score.to_json()

            context = {
                'opt_plot': request.session.get('opt plot'),
                'opt_score_table': request.session.get('opt score table'),
                'ra_plot': ra_fig_div,
                'ra_score_table': ra_score_table,
                'plot_vhr': plot_vhr,
                'plot_adhesion': plot_adhesion,
                'plot_lts': plot_lts,
                # 'plot_lto': plot_lto,
                'plot_delta_angle': plot_delta_angle,
                'plot_pct': plot_pct,
                'plot_sealwvtr': plot_sealwvtr,
            }

            return render(request, 'filteredResult.html', context=context)
    return redirect(reverse('index'))


def xlsx_export(request):

    df = pd.read_json(request.session['ra_result'])
    mean_df = pd.read_json(request.session['ra_score'])

    with BytesIO() as b:
        # Use the StringIO object as the filehandle
        writer = pd.ExcelWriter(b, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='RA Result', index=False)
        mean_df.to_excel(writer, sheet_name='RA Score', index=False)
        if request.session.get('opt result') and request.session.get('opt score'):
            opt_result = pd.read_json(request.session['opt result'])
            opt_score = pd.read_json(request.session['opt score'])
            opt_result.to_excel(
                writer, sheet_name='OPT Result', index=False)
            opt_score.to_excel(
                writer, sheet_name='OPT Score', index=False)
            mean_df.to_excel(
                writer, sheet_name='RA Score', index=False)
            opt_score = opt_score.rename(columns={'sum': 'opt score'})
            # total_score = ra_score[['LC', 'PI', 'Seal', 'ra score']].merge(
            #     opt_score[['LC', 'opt score']], on='LC', how='left')
            # total_score['score'] = total_score['ra score'] + \
            #     total_score['opt score']
            # total_score = total_score.sort_values(by='score', ascending=False)
            # total_score.to_excel(writer, sheet_name='Total Score', index=False)
        writer.save()
        # Set up the HTTP response
        filename = 'Result.xlsx'
        response = HttpResponse(
            b.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response


def test(request):
    VHR_query = VHR.objects.filter(
        value__gt=50).order_by('vender__name', '-value')
    x_data = []
    y_data = []
    vender = []
    for row in VHR_query:
        x_data += [f'{row.LC.name} {row.PI.name} {row.seal.name}']
        vender += [row.vender.name]
        y_data += [float(row.value)]
        # print(f'{row.value}\t{row.LC.name} {row.PI.name} {row.seal.name}')

    df = pd.DataFrame({
        'cond': x_data,
        'VHR(%)': y_data,
        'vender': vender,
    })

    df_mean = df.groupby(by=['vender', 'cond'], as_index=False).mean()

    fig = px.bar(df_mean, x='cond', y='VHR(%)',
                 color='vender', barmode='group')
    plot_div = plot(fig, output_type='div')
    request.session['df'] = df.to_json()

    return render(request, 'test.html', context={'plot_div': plot_div})


def test_download(request):
    df = pd.read_json(request.session['df'])
    with BytesIO() as b:
        # Use the StringIO object as the filehandle.
        writer = pd.ExcelWriter(b, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1')
        writer.save()
        # Set up the Http response.
        filename = 'django_simple.xlsx'
        response = HttpResponse(
            b.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response


class BatchUploadView(View):
    def get(self, request, *args, **kwargs):
        return render(
            request,
            'batchUpload.html',
            context=({
                'file_form': UploadFileForm,
            })
        )
